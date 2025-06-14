import numpy as np
import pygame
import pyvisgraph as vg
from shapely.geometry import Point, Polygon
import matplotlib.pyplot as plt
from openpyxl import load_workbook, Workbook
import os

# Simulation parameters 
## ALL SPATIAL VALUES ARE DOUBLED FOR THE SAKE OF VISUALISATION ##
times_to_run_simulation = 50 # remove output.xlsx before running if you want to save the data (if there is one)
time_step = 0.1 # Unit is seconds
number_of_agents = 30
agent_radius = 5 # Unit is dm (0.1 m), source https://dined.io.tudelft.nl/en/database/tool "Breath over elbows" (simplified to 5) (In reality 2.5)
agent_max_speed = 12.7*2 # max moving speed, unit is dm/s (0.1 m/s), source SFPE handbook of FPE Table 3-13.5 (doubled because spatial distances are doubled too)
sd_reaction_time = 30 
avg_reaction_time = 6*60 + sd_reaction_time 
# average time in seconds for agents to react to fire alarm, only the variance considered and
# average just added to the total time from SFPE handbook of FPE Table 3-13.1 (6 minutes) so here the reaction time and variance are same even if in reality thet aren't


## Building the building blueprint ##

# Exit location(s)
exit_locations = [[330,600]]

# Walls
# top left corner is (0.0, 0.0) and down right corner is (SCREEN_WIDTH, SCREEN_HEIGHT), (x,y)
# here unit is also dm but double here too all values
polys = [[vg.Point(100,700), vg.Point(300,700), vg.Point(300, 614), vg.Point(305,614),
          vg.Point(305, 705), vg.Point(95,705), vg.Point(95,495), vg.Point(305, 495),
          vg.Point(305, 586), vg.Point(300, 586), vg.Point(300, 500), vg.Point(100, 500)]]
polys[0].reverse()

# initialization for visualization
SCREEN_WIDTH = 1200
SCREEN_HEIGHT = 800

# Colour initialization for pygame
RED = (255, 0, 0)
GREEN = (0, 255, 0)
BLUE = (0, 0, 255)
YELLOW = (255, 255, 0)
WHITE = (255, 255, 255)
BLACK = (0, 0, 0)
PURPLE = (128, 0, 128)
ORANGE = (255, 165 ,0)
GREY = (128, 128, 128)
TURQUOISE = (64, 224, 208)

class Exit:
    def __init__(self, x, y):
        self.x = x
        self.y = y

class Agent:
    def __init__(self, x, y, vx, vy, speed, radius, exit, reaction_time):
        self.x = x                          # Current x position
        self.y = y                          # Current y position
        self.vx = vx                        # Velocity in x direction
        self.vy = vy                        # Velocity in y direction
        self.speed = speed                  # Current speed
        self.radius = radius                # Radius
        self.targetx = exit.x               # Location of exit x-coordinate
        self.targety = exit.y               # Location of exit y-coordinate
        self.path = []                      # Path in form of visuality graph and all the nodes in this list
        self.path_index = 0                 # Index for the next point in the path
        self.safe = False                   # Indicates whether agent has safely exited the building
        self.reset_path = False             # Shows that path needs to be calculated again
        self.waiting = True                 # Waits for other agent to pass
        self.reaction_time = reaction_time + avg_reaction_time  # Time to react to fire alarm

    def draw(self, screen):
        pygame.draw.circle(screen, RED, (int(self.x), int(self.y)), self.radius)
    def update(self, dt):
        self.x += self.vx * dt
        self.y += self.vy * dt

    ## Moving ##
    def checkClosestRoute(self):
        # Takes the agent radius into consideration
        # 5 added to radius to make the agent walk a bit more away from the walls
        inflated_obstacles = [inflatePolygon(polygon, self.radius + 5) for polygon in polys]

        g = vg.VisGraph()
        g.build(inflated_obstacles)

        start = vg.Point(self.x, self.y)
        goal = vg.Point(self.targetx, self.targety)

        self.path = g.shortest_path(start, goal)
        self.path_index = 0
    def moveAlongPath(self, dt):
        if self.path_index < len(self.path):
            # Target the next step on the path
            target_point = self.path[self.path_index]
            target_x, target_y = target_point.x, target_point.y

            self.moveTowards(target_x, target_y, dt) # Move the agent

            # Checks if agent is close enough for next step
            if np.sqrt((self.x - target_x) ** 2 + (self.y - target_y) ** 2) < 1 and self.reset_path:
                self.checkClosestRoute()
                self.path_index += 1
                self.reset_path = False
            elif np.sqrt((self.x - target_x) ** 2 + (self.y - target_y) ** 2) < 1:
                self.path_index += 1

            # Checks if agent is close to the target
            if np.sqrt((self.x - self.targetx) ** 2 + (self.y - self.targety) ** 2) < 1:
                self.safe = True
    def moveTowards(self, target_x, target_y, dt):
        ## Direction ##
        direction_x = target_x - self.x
        direction_y = target_y - self.y
        distance = np.sqrt(direction_x**2 + direction_y**2)

        ## Speed ##
        # Just for preventing any crazy speeds
        if self.speed > agent_max_speed:
            self.speed = agent_max_speed
        # Try to always keep the maximum speed
        if self.speed < agent_max_speed:
            self.speed *= 1.05
        if self.speed < 1:
            self.speed = 1

        # Slow down when near next path step i.e in corners
        if distance < 25:
            if self.speed*1.5 > agent_max_speed:
                self.speed *= 0.92
                
        if distance > 0:
            self.vx = (direction_x / distance) * self.speed
            self.vy = (direction_y / distance) * self.speed
            #print(f"({self.vx},{self.vy})")

        self.update(dt) 
    def dodge(self, distances):
        sector_range = 50
        # Current direction
        target_point = self.path[self.path_index]
        direction_x = target_point.x - self.x
        direction_y = target_point.y - self.y
        distance = np.sqrt(direction_x**2 + direction_y**2)
        if distance == 0:
            distance = 1
        rng_vector = np.random.rand(2)
        if (distances[0]*2)/sector_range < rng_vector[0]: # dodging agents
            if rng_vector[1] < 0.5: # dodge to left
                new_direction_x = -direction_y/distance
                new_direction_y = direction_x/distance
                if distances[1] < distances[4]:
                    new_x = self.x + new_direction_x * (distances[1] - self.radius - 6)
                    new_y = self.y + new_direction_y * (distances[1] - self.radius - 6)
                else:
                    new_x = self.x + new_direction_x * (distances[4] - self.radius - 6)
                    new_y = self.y + new_direction_y * (distances[4] - self.radius - 6)
                new_path_point = vg.Point(new_x, new_y)
                self.path.insert(self.path_index, new_path_point)
            if rng_vector[1] > 0.5: # dodge to right
                new_direction_x = direction_y/distance
                new_direction_y = -direction_x/distance
                if distances[2] < distances[5]:
                    new_x = self.x + new_direction_x * (distances[2] - self.radius - 6)
                    new_y = self.y + new_direction_y * (distances[2] - self.radius - 6)
                else:
                    new_x = self.x + new_direction_x * (distances[5] - self.radius - 6)
                    new_y = self.y + new_direction_y * (distances[5] - self.radius - 6)
                new_path_point = vg.Point(new_x, new_y)
                self.path.insert(self.path_index, new_path_point)
            self.reset_path = True
    def randomMovement(self, distances):
        # Current direction
        target_point = self.path[self.path_index]
        direction_x = target_point.x - self.x
        direction_y = target_point.y - self.y
        distance = np.sqrt(direction_x**2 + direction_y**2)
        if distance == 0:
            distance = 1
        rng_vector = np.random.rand(2)
        if 0.002 > rng_vector[0]:
            if rng_vector[1] < 0.5: # random movement to left
                new_direction_x = -direction_y/distance
                new_direction_y = direction_x/distance
                if distances[1] < distances[4]:
                    new_x = self.x + new_direction_x * (distances[1] - self.radius - 6)
                    new_y = self.y + new_direction_y * (distances[1] - self.radius - 6)
                else:
                    random_movement_distance = distances[4]*0.5 + np.random.rand()*(distances[4]*0.5)
                    new_x = self.x + new_direction_x * (random_movement_distance - self.radius - 6)
                    new_y = self.y + new_direction_y * (random_movement_distance - self.radius - 6)
                new_path_point = vg.Point(new_x, new_y)
                self.path.insert(self.path_index, new_path_point)
            if rng_vector[1] > 0.5: # random movement to right
                new_direction_x = direction_y/distance
                new_direction_y = -direction_x/distance
                if distances[2] < distances[5]:
                    new_x = self.x + new_direction_x * (distances[2] - self.radius - 6)
                    new_y = self.y + new_direction_y * (distances[2] - self.radius - 6)
                else:
                    random_movement_distance = distances[5]*0.5 + np.random.rand()*(distances[5]*0.5)
                    new_x = self.x + new_direction_x * (random_movement_distance - self.radius - 6)
                    new_y = self.y + new_direction_y * (random_movement_distance - self.radius - 6)
                new_path_point = vg.Point(new_x, new_y)
                self.path.insert(self.path_index, new_path_point)
            self.reset_path = True
        return None 

    ## Collision detection and handling ##
    def detectCollision(self, agents, dt):
        # Draw the circle where the agent would be in the next time step (*2 just to give some space)
        agent_circle = Point(self.x + self.vx*dt*2, self.y + self.vy*dt*2).buffer(self.radius)

        # Collision with polygon check
        for poly in polys:
            shapely_poly = Polygon([(p.x, p.y) for p in poly]) # Turn pyvisgraph poly into shapely poly
            if agent_circle.intersects(shapely_poly):
                return "wall"
            
        # Collision with other agents check
        for other_agent in agents:
            if other_agent == self:
                continue

            other_agent_circle = Point(other_agent.x + other_agent.vx*dt, other_agent.y + other_agent.vy*dt).buffer(other_agent.radius)
            collision = agent_circle.intersects(other_agent_circle)

            # If no collision
            if not collision:
                continue

            # Prioritize faster moving agent
            if collision and not self.waiting and not other_agent.waiting:
                if self.speed < other_agent.speed:
                    self.waiting = True
                else:
                    other_agent.waiting = True

            # If both are waiting for some reason
            elif collision and self.waiting and other_agent.waiting:
                if np.random.rand() > 0.5:
                    self.waiting = False
                else:
                    other_agent.waiting = False

            # If agent is waiting
            if collision and self.waiting and not other_agent.waiting:
                return "waiting"
            
            # If agent is passing
            elif collision and not self.waiting and other_agent.waiting:
                return "passing"
            
        # If no collision is detected
        return False
    def objectInSector(self, agents):
        sector_polygon, sector_right, sector_left = self.createSectorPolygon()
        sector_range = 50
        distances = [50, 25, 25, 50, 25, 25] # Distances to agents in front(0), left(1) and right(2) and distances to walls in front(3), left(4) and right(5)

        # Seeing other agents check
        for other_agent in agents:
            if other_agent == self:
                continue 

            other_agent_point = Point(other_agent.x, other_agent.y)
            

            # Find the distance to the closest agent 
            if sector_polygon.intersects(other_agent_point):
                distance_to_agent = np.sqrt((self.x - other_agent.x) ** 2 + (self.y - other_agent.y) ** 2)
                if distances[0] > distance_to_agent:
                    distances[0] = distance_to_agent
            # Find the distance to the closest agent in left sector
            if sector_left.intersects(other_agent_point):
                distance_to_agent = np.sqrt((self.x - other_agent.x) ** 2 + (self.y - other_agent.y) ** 2)
                if distances[1] > distance_to_agent:
                    distances[1] = distance_to_agent
            # Find the distance to the closest agent in right sector
            if sector_right.intersects(other_agent_point):
                distance_to_agent = np.sqrt((self.x - other_agent.x) ** 2 + (self.y - other_agent.y) ** 2)
                if distances[2] > distance_to_agent:
                    distances[2] = distance_to_agent
            # Wall checks
            for poly in polys:
                shapely_poly = Polygon([(p.x, p.y) for p in poly]) # Turn pyvisgraph poly into shapely poly
                # Find the distance to the closest wall
                if sector_polygon.intersects(shapely_poly):
                    distance_to_wall = shapely_poly.distance(Point(self.x, self.y))
                    if distances[3] > distance_to_wall:
                        distances[3] = distance_to_wall
                # Find the distance to the closest wall in left sector
                if sector_left.intersects(shapely_poly):
                    distance_to_wall = shapely_poly.distance(Point(self.x, self.y))
                    if distances[4] > distance_to_wall:
                        distances[4] = distance_to_wall
                # Find the distance to the closest wall in right sector
                if sector_right.intersects(shapely_poly):
                    distance_to_wall = shapely_poly.distance(Point(self.x, self.y))
                    if distances[5] > distance_to_wall:
                        distances[5] = distance_to_wall
        # Slow down the agent depending on the distance to the other agents in all sectors
        if distances[0] < sector_range:
            self.speed = agent_max_speed * ((distances[0] - sector_range - agent_radius*2 - 5)/(sector_range)) + agent_max_speed
            if self.speed < 0:
                self.speed = 0
        return distances
    def createSectorPolygon(self):
        angle_rad = np.radians(45 / 2)
        angle_rad_2 = angle_rad * 4
        range_dist = 50
        range_dist_2 = range_dist/2

        points = [Point(self.x, self.y)]
        points_left = [Point(self.x, self.y)]
        points_right = [Point(self.x, self.y)]

        # 1st sector
        for angle in np.linspace(-angle_rad, angle_rad, 5):
            x = self.x + range_dist * np.cos(np.atan2(self.vy, self.vx) + angle)
            y = self.y + range_dist * np.sin(np.atan2(self.vy, self.vx) + angle)
            points.append(Point(x, y))
        
        # left sector
        for angle in np.linspace(-angle_rad_2, -angle_rad, 5):
            x = self.x + range_dist_2 * np.cos(np.atan2(self.vy, self.vx) + angle)
            y = self.y + range_dist_2 * np.sin(np.atan2(self.vy, self.vx) + angle)
            points_left.append(Point(x, y))

        # right sector
        for angle in np.linspace(angle_rad, angle_rad_2, 5):
            x = self.x + range_dist_2 * np.cos(np.atan2(self.vy, self.vx) + angle)
            y = self.y + range_dist_2 * np.sin(np.atan2(self.vy, self.vx) + angle)
            points_right.append(Point(x, y))

        points.append(Point(self.x, self.y))
        points_left.append(Point(self.x, self.y))
        points_right.append(Point(self.x, self.y))

        sector_polygon = Polygon([(p.x, p.y) for p in points])
        #sector_coords = [(int(p.x), int(p.y)) for p in points]
        #sector_coords_left = [(int(p.x), int(p.y)) for p in points_left]
        sector_polygon_left = Polygon([(p.x, p.y) for p in points_left])
        #sector_coords_right = [(int(p.x), int(p.y)) for p in points_right]
        sector_polygon_right = Polygon([(p.x, p.y) for p in points_right])
        #pygame.draw.polygon(screen, BLUE, sector_coords, 1)
        #pygame.draw.polygon(screen, RED, sector_coords_left, 1)
        #pygame.draw.polygon(screen, BLUE, sector_coords_right, 1)

        return sector_polygon, sector_polygon_left, sector_polygon_right
    def moveAwayFromWall(self, agents, dt):
        start = np.pi
        end = start + np.pi*2
        for angle in np.linspace(start, end, 10):
            x_direction = np.cos(angle)
            y_direction = np.sin(angle)
            distance = np.sqrt(x_direction**2 + y_direction**2)
            if distance > 0:
                self.vx = (x_direction / distance) * agent_max_speed
                self.vy = (y_direction / distance) * agent_max_speed
            # Time step multiplied by 5 to get the agent far away from the wall
            if self.detectCollision(agents, dt*5) != "wall":
                print("Moving away from a wall")
                self.update(dt*5)
                return None

## Spawn agents ##

def spawnAgents(exits):
    agents = []
    for _ in range(number_of_agents):
        while True:
            x = np.random.uniform(100, 300)
            y = np.random.uniform(500, 700)

            if not validPosition((x, y), agents):
                speed = 0
                exit = np.random.choice(exits)
                reaction_time = np.random.normal(sd_reaction_time, sd_reaction_time)
                radius = agent_radius
                agent = Agent(x, y, 0, 0, speed, radius, exit, reaction_time)
                agents.append(agent)
                break
    return agents
def validPosition(point, agents):
    x, y = point
    agent_circle = Point(x, y).buffer(agent_radius+5)
    
    for poly in polys:
        shapely_poly = Polygon([(p.x, p.y) for p in poly])
        if shapely_poly.intersects(agent_circle):
            return True
    for other_agent in agents:
        other_agent_circle = Point(other_agent.x + other_agent.vx, other_agent.y + other_agent.vy).buffer(other_agent.radius)
        if agent_circle.intersects(other_agent_circle):
            return True
    return False

## Generate exits ##

def generateExits():
    exits = []
    for location in exit_locations:
        exit = Exit(location[0], location[1])
        exits.append(exit)
    return exits

## Inflating the polygons for pathfinding ##

def normalize(vector):
    length = np.linalg.norm(vector)
    return vector / length if length != 0 else vector
def inflatePolygon(polygon, radius):
    inflated_polygon = []
    num_vertices = len(polygon)
    previous_normal_vector = np.array([])
    
    for i in range(num_vertices):
        # Current edge's start and end points
        p1 = np.array([polygon[i].x, polygon[i].y])
        p2 = np.array([polygon[(i + 1) % num_vertices].x, polygon[(i + 1) % num_vertices].y])
        
        # Edge vector and perpendicular normal vector
        edge_vector = p2 - p1
        normal_vector = np.array([-edge_vector[1], edge_vector[0]])
        normal_vector = normalize(normal_vector) * radius

        if previous_normal_vector.size != 0:
            if isInnerCorner(previous_normal_vector, normal_vector):
                inflated_p2 = p2 + normal_vector
                inflated_polygon.append(vg.Point(inflated_p2[0], inflated_p2[1]))
            else:
                inflated_p1 = p1 + normal_vector
                inflated_p2 = p2 + normal_vector
                inflated_polygon.append(vg.Point(inflated_p1[0], inflated_p1[1]))
                inflated_polygon.append(vg.Point(inflated_p2[0], inflated_p2[1]))
            previous_normal_vector = normal_vector
        else:
            previous_normal_vector = normal_vector
            inflated_p1 = p1 + normal_vector
            inflated_p2 = p2 + normal_vector
            inflated_polygon.append(vg.Point(inflated_p1[0], inflated_p1[1]))
            inflated_polygon.append(vg.Point(inflated_p2[0], inflated_p2[1]))
    
    return inflated_polygon
def isInnerCorner(previous_normal_vector, current_normal_vector):
    dot_product = previous_normal_vector[0] * current_normal_vector[0] + previous_normal_vector[1] * current_normal_vector[1]
    cross_product = previous_normal_vector[0] * current_normal_vector[1] - previous_normal_vector[1] * current_normal_vector[0]
    # Inner corner happens when:
    # - Dot product is negative (angle >= 90 degrees)
    # - Cross product is positive
    if dot_product <= 0 and cross_product > 0:
        return True  # inner corner
    else:
        return False  # not an inner corner

## Plot ##

def plotTimes(safe_times):
    x = safe_times
    y = np.arange(1, len(safe_times) + 1)

    plt.figure(figsize=(8, 5))
    plt.step(x, y, where='post', label='Cumulative exits')
    plt.xlabel('Time (s)')
    plt.ylabel('Number of agents evacuated')
    plt.title('Agent exits over time')
    plt.grid(True)
    plt.legend()
    plt.show()

## Write an excel ##

def writeExcel(safe_times, iteration):
    filename = "output.xlsx"

    if iteration == 0:
        workbook = Workbook()
        worksheet = workbook.active
    if os.path.exists(filename):
        workbook = load_workbook(filename)
        worksheet = workbook.active
    

    worksheet.cell(1, iteration+1, f"Iteration {iteration+1}")
    for i, time in enumerate(safe_times):
        worksheet.cell(i+2, iteration+1).value = time


    workbook.save(filename)



## Running of the code ##

def main(iteration):
    pygame.init()
    screen = pygame.display.set_mode((SCREEN_WIDTH, SCREEN_HEIGHT))

    run = True
    exits = generateExits()
    agents = spawnAgents(exits)
    clock = pygame.time.Clock() # Used only for visualisation
    time = avg_reaction_time
    safe_times = []

    # Calculates initial paths
    for agent in agents:
        agent.checkClosestRoute()

    while run:
        for event in pygame.event.get():
            if event.type == pygame.QUIT:
                run = False
        
        # Time in seconds
        time += time_step
        
        # Clears the screen between each time step
        screen.fill(WHITE)

        # Agent movement logic
        for i in range(len(agents)-1, -1, -1): # backwards iteration to pop agents during the loop
            agent = agents[i]
            distances = agent.objectInSector(agents)
            collision = agent.detectCollision(agents, time_step)
            if not collision and not agent.waiting:
                if distances[0] < 50:
                    agent.dodge(distances)
                else:
                    agent.randomMovement(distances)
                agent.moveAlongPath(time_step)
            elif collision == "wall":
                agent.speed = 0
                agent.moveAwayFromWall(agents, time_step)
                agent.checkClosestRoute()
            elif collision == "waiting":
                agent.speed = 0
            elif collision == "passing":
                agent.moveAlongPath(time_step)
            elif agent.reaction_time <= time:
                agent.waiting = False
            agent.draw(screen)
            if agent.safe == True:
                safe_times.append(time)
                agents.pop(i)
        
        if len(agents) == 0:
            #plotTimes(safe_times)
            writeExcel(safe_times, iteration)
            run = False

        # Draws obstacles
        for poly in polys:
            polygon_points = [(int(p.x), int(p.y)) for p in poly]
            pygame.draw.polygon(screen, BLACK, polygon_points, 0)
   
        pygame.display.update()

        # Caps the frame rate to 60 FPS
        clock.tick(60)

    pygame.quit()

# Run simulation x times

for i in range(times_to_run_simulation):
    main(i)